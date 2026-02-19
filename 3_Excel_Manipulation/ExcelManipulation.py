from openpyxl import Workbook, load_workbook

# --- Create a new Excel file ---
wb = Workbook()
ws = wb.active
ws.title = "SampleSheet"

# Add some data
ws['A1'] = "Name"
ws['B1'] = "Age"

ws.append(["Alice", 25])
ws.append(["Bob", 30])
ws.append(["Charlie", 28])

# Save the file
wb.save("sample.xlsx")
print("Excel file 'sample.xlsx' created successfully!")

# --- Read the Excel file ---
wb2 = load_workbook("sample.xlsx")
sheet = wb2["SampleSheet"]

print("\nReading from Excel:")
for row in sheet.iter_rows(values_only=True):
    print(row)
